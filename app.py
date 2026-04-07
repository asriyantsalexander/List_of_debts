import re
import os
import io
import tempfile
import subprocess
from flask import Flask, request, send_file, render_template, jsonify
from pypdf import PdfReader
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
 
app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 200 * 1024 * 1024  # 200 MB
 
# ── Extraction logic ────────────────────────────────────────────────────────
 
def last_day(d):
    _, m, y = map(int, d.split('.')); m -= 1
    if m == 0: m = 12; y -= 1
    ends = {1:31,2:28,3:31,4:30,5:31,6:30,7:31,8:31,9:30,10:31,11:30,12:31}
    if m == 2 and (y%4==0 and (y%100!=0 or y%400==0)): ends[2] = 29
    return f"{ends[m]:02d}.{m:02d}.{y}"
 
def prev_1st(d):
    _, m, y = map(int, d.split('.')); m -= 1
    if m == 0: m = 12; y -= 1
    return f"01.{m:02d}.{y}"
 
def prev_11th(d):
    _, m, y = map(int, d.split('.')); m -= 1
    if m == 0: m = 12; y -= 1
    return f"11.{m:02d}.{y}"
 
def extract_card(pdf_path):
    r = PdfReader(pdf_path)
    text = ''.join(p.extract_text() for p in r.pages) + '\n'
    hdr  = text[:800]
 
    otp_count   = hdr.count('Отопление')
    has_pov     = 'повышающий' in hdr
    has_en      = 'Горячее в/с (энергия)' in hdr
    has_nos     = 'Горячее в/с (носитель)' in hdr
    has_otp     = 'Отопление' in hdr
    triple_otp  = otp_count >= 3
    dual_otp    = otp_count >= 2
    gvs_first   = (has_en and has_otp and
                   hdr.find('Горячее в/с (энергия)') < hdr.find('Отопление'))
    en_before_nos = (has_en and has_nos and
                     hdr.find('Горячее в/с (энергия)') < hdr.find('Горячее в/с (носитель)'))
 
    vp = r'\s+([-\d.]+)'
    all_saldo, n_cols = [], 1
    for nc in [6, 5, 4, 3, 2, 1]:
        found = re.findall(r'Сальдо на (\d{2}\.\d{2}\.\d{4})' + vp*(nc+1), text)
        if found:
            all_saldo, n_cols = found, nc
            break
    if not all_saldo:
        return None
 
    def parse(entry):
        v   = [float(x) for x in entry[1:]]
        svc = v[:-1]   # drop ИТОГО
 
        # Triple Отопление: ОТП1 | ГВС_эн | ГВС_нос | ГВС_нос_пов | ОТП2 | ОТП3 | ИТОГО
        if triple_otp and n_cols == 6 and en_before_nos and not gvs_first:
            otp1, en, nos, pov, otp2, otp3 = svc
            return otp1 + otp2 + otp3, en, nos + pov
 
        # gvs_first: ГВС_эн | ОТП | ГВС_нос | ГВС_нос_пов | ГВС_нос2 | ИТОГО
        if gvs_first and n_cols == 5:
            en, otp, nos, pov, nos2 = svc
            return otp, en, nos + pov + nos2
 
        # Dual Отопление + повышающий: ОТП1 | ГВС_нос | ГВС_эн | ГВС_нос_пов | ОТП2 | ИТОГО
        if dual_otp and has_pov and n_cols == 5:
            otp1, nos, en, pov, otp2 = svc
            return otp1 + otp2, en, nos + pov
 
        # Dual Отопление without повышающий: ОТП1 | ГВС_нос | ГВС_эн | ОТП2 | ИТОГО
        if dual_otp and not has_pov and n_cols == 4:
            otp1, nos, en, otp2 = svc
            return otp1 + otp2, en, nos
 
        # Standard 4 services with повышающий: ОТП | ГВС_нос | ГВС_эн | ГВС_нос_пов | ИТОГО
        if has_pov and n_cols == 4:
            otp, nos, en, pov = svc
            return otp, en, nos + pov
 
        # Standard 3 services: ОТП | ГВС_нос | ГВС_эн | ИТОГО
        if n_cols == 3:
            otp, nos, en = svc
            return otp, en, nos
 
        # Single service: ОТП | ИТОГО
        return svc[0], 0.0, 0.0
 
    last_date = all_saldo[-1][0]
    end_date  = last_day(last_date)
    lv = parse(all_saldo[-1])
    td, pd, ted = (max(0.0, x) for x in lv)
 
    # First POSITIVE saldo per service
    fps = [None, None, None]
    for e in all_saldo:
        t, p, te = parse(e)
        dt = e[0]
        if fps[0] is None and t  > 0: fps[0] = dt
        if fps[1] is None and p  > 0: fps[1] = dt
        if fps[2] is None and te > 0: fps[2] = dt
    first_d = [prev_1st(x) if x else None for x in fps]
 
    # Last peni saldo (remaining peni debt, not total charged)
    peni_pat = r'Сальдо пени на конец периода' + vp*(n_cols+1)
    all_peni = re.findall(peni_pat, text)
    if all_peni:
        lp = parse(['_'] + list(all_peni[-1]))
        tp, pp, tep_p = (max(0.0, x) for x in lp)
    else:
        tp = pp = tep_p = 0.0
 
    # First positive peni per service — position-based scan (robust across environments)
    fpp = [None, None, None]
    all_peni_matches  = list(re.finditer(r'Сальдо пени на конец периода' + vp*(n_cols+1), text))
    all_saldo_matches = list(re.finditer(r'Сальдо на (\d{2}\.\d{2}\.\d{4})', text))
    for pm in all_peni_matches:
        next_sd = next((m.group(1) for m in all_saldo_matches if m.start() > pm.start()), None)
        if not next_sd: continue
        peni_start = prev_11th(next_sd)
        pv = parse(['_'] + list(pm.groups()))
        if fpp[0] is None and pv[0] > 0: fpp[0] = peni_start
        if fpp[1] is None and pv[1] > 0: fpp[1] = peni_start
        if fpp[2] is None and pv[2] > 0: fpp[2] = peni_start
        if all(x is not None for x in fpp): break
 
    return dict(
        teplo_dolg=round(td, 2),  pod_dolg=round(pd, 2),   tep_dolg=round(ted, 2),
        teplo_peni=round(tp, 2),  pod_peni=round(pp, 2),   tep_peni=round(tep_p, 2),
        first_d=first_d, first_p=fpp, end_date=end_date
    )
 
 
def build_xlsx(results, acc_order):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Лист2'
 
    thin = Side(style='thin', color='000000')
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
    h9   = Font(name='Calibri', bold=True, size=9)
    d11  = Font(name='Calibri', size=11)
    ca   = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cc   = Alignment(horizontal='center', vertical='center')
 
    def hc(c, v): c.value=v; c.font=h9;  c.alignment=ca; c.border=bdr
    def dc(c, v): c.value=v; c.font=d11; c.alignment=cc; c.border=bdr
 
    ws.merge_cells('C2:F2'); ws.merge_cells('G2:J2'); ws.merge_cells('K2:N2')
    hc(ws['A2'], None); hc(ws['B2'], None)
    hc(ws['C2'], 'Теплоснабжение')
    hc(ws['G2'], 'Горячее водоснабжение - подогрев')
    hc(ws['K2'], 'Горячее водоснабжение - теплоноситель')
 
    for col, h in enumerate([
        'Лицевой счет', 'Назначение платежа',
        'Задолженность', 'Период', 'Пени', 'Период',
        'Задолженность', 'Период', 'Пени', 'Период',
        'Задолженность', 'Период', 'Пени', 'Период'], 1):
        hc(ws.cell(3, col), h)
 
    for col, w in zip('ABCDEFGHIJKLMN',
                      [12, 40, 12, 22, 10, 22, 12, 22, 10, 22, 12, 22, 10, 22]):
        ws.column_dimensions[col].width = w
 
    def ps(start, val, end):
        return f"{start}\u2013{end}" if (start and val > 0) else None
 
    for row_i, ls in enumerate(acc_order, start=4):
        if ls not in results: continue
        d   = results[ls]
        end = d['end_date']
        fd  = d['first_d']
        fp  = d['first_p']
        for col, v in enumerate([
            int(ls),
            f'взыскание задолженности лс {ls}',
            d['teplo_dolg'], ps(fd[0], d['teplo_dolg'], end),
            d['teplo_peni'], ps(fp[0], d['teplo_peni'], end),
            d['pod_dolg'],   ps(fd[1], d['pod_dolg'],   end),
            d['pod_peni'],   ps(fp[1], d['pod_peni'],   end),
            d['tep_dolg'],   ps(fd[2], d['tep_dolg'],   end),
            d['tep_peni'],   ps(fp[2], d['tep_peni'],   end),
        ], 1):
            dc(ws.cell(row_i, col), v)
 
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
 
 
# ── Routes ──────────────────────────────────────────────────────────────────
 
@app.route('/')
def index():
    return render_template('index.html')
 
 
@app.route('/process', methods=['POST'])
def process():
    if 'archive' not in request.files:
        return jsonify({'error': 'Файл не загружен'}), 400
 
    f = request.files['archive']
    if not f.filename:
        return jsonify({'error': 'Файл не выбран'}), 400
 
    with tempfile.TemporaryDirectory() as tmp:
        archive_path = os.path.join(tmp, f.filename)
        f.save(archive_path)
 
        extract_dir = os.path.join(tmp, 'cards')
        os.makedirs(extract_dir)
 
        fname_lower = f.filename.lower()
        try:
            if fname_lower.endswith('.rar'):
                subprocess.run(['unrar', 'x', archive_path, extract_dir],
                               check=True, capture_output=True)
            elif fname_lower.endswith('.zip'):
                subprocess.run(['unzip', '-o', archive_path, '-d', extract_dir],
                               check=True, capture_output=True)
            elif fname_lower.endswith('.7z'):
                subprocess.run(['7z', 'x', archive_path, f'-o{extract_dir}'],
                               check=True, capture_output=True)
            else:
                return jsonify({'error': 'Поддерживаются форматы: .rar, .zip, .7z'}), 400
        except subprocess.CalledProcessError as e:
            return jsonify({'error': f'Ошибка распаковки: {e.stderr.decode()}'}), 500
 
        pdfs = []
        for root, dirs, files in os.walk(extract_dir):
            for fname in files:
                if fname.lower().endswith('.pdf'):
                    pdfs.append((fname.replace('.pdf', '').replace('.PDF', ''),
                                 os.path.join(root, fname)))
        pdfs.sort(key=lambda x: x[0])
 
        if not pdfs:
            return jsonify({'error': 'PDF-файлы не найдены в архиве'}), 400
 
        results, acc_order, errors = {}, [], []
        for ls, path in pdfs:
            try:
                d = extract_card(path)
                if d:
                    results[ls] = d
                    acc_order.append(ls)
                else:
                    errors.append(ls)
            except Exception as e:
                errors.append(f'{ls} ({e})')
 
        if not results:
            return jsonify({'error': 'Не удалось обработать ни один файл'}), 500
 
        buf = build_xlsx(results, acc_order)
 
    return send_file(
        buf,
        as_attachment=True,
        download_name='задолженность.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
 
 
if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port)
 
